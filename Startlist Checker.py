import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import re

# define the base URL
base_url = 'https://www.procyclingstats.com/'

# specify the year to check
year = '2023'

# load the workbook
wb = load_workbook('riders.xlsx')
races_sheet = wb['races']
riders_sheet = wb['riders']

## get the Startlist sheet or create a new one if it doesn't exist
if 'Startlist' in wb.sheetnames:
    Startlist_sheet = wb['Startlist']
    Startlist_sheet.delete_rows(2, Startlist_sheet.max_row)
else:
    Startlist_sheet = wb.create_sheet('Startlist')

# get the race names and numbers from the races sheet
race_names = []
race_numbers = []
for row in races_sheet.iter_rows(min_row=2, min_col=1, values_only=True):
    race_name = row[0]
    race_number = row[1]
    race_rating = row[2]
    race_names.append(race_name)
    race_numbers.append(race_number)

# sort the race names according to the race numbers
sorted_race_names = [x for _,x in sorted(zip(race_numbers, race_names))]

# write the rider names and race names to the Startlist sheet
Startlist_sheet['A1'] = 'Rider'
Startlist_sheet['B1'] = '# races'
riders_sheet['D1'] = '# races'
for i, race_name in enumerate(sorted_race_names):
    Startlist_sheet.cell(row=1, column=i+3, value=race_name)

# initialize row_num to 2
row_num = 2

# loop through each race in the races sheet and build the startlist
for i, race_name in enumerate(sorted_race_names):
        # extract the race name and build the startlist URL
        startlist_url = base_url + 'race/{}/'.format(str(sorted_race_names[i]).lower()) + year + '/startlist'

        # get the HTML content of the startlist URL
        response = requests.get(startlist_url)
        html_content = response.text

        # parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')

        # Find the startlist_v3 element in the HTML content
        startlist_v3 = soup.find('ul', class_='startlist_v3')

        # check if the startlist table was found
        if startlist_v3 is None:
            print('Startlist table not found for race', race_name)
            continue

        # loop through each rider in the riders sheet and check if they're in the startlist
        for j, rider in enumerate(riders_sheet.iter_rows(min_row=2, values_only=True)):
            
            # check if the rider name is not None
            if rider[1] is None:
                continue
            
            # extract the rider name, team and value and replace special characters with regular ones
            rider_name = re.sub(r'[^\w\s]', '', rider[1]).lower().replace('.', '')
            rider_team = rider[0].lower()
            rider_value = rider[2]

            # initialize the dictionary to hold the results for the rider
            rider_results = {}          
                     
            # check if the rider name is in the startlist table
            if re.sub(r'[^a-zA-Z\s]', '', rider_name) in re.sub(r'[^a-zA-Z\s]', '', startlist_v3.text.lower()):
                rider_results[race_name] = 1                         
            else:
                rider_results[race_name] = 0

            # write the results for the rider to the Startlist sheet
            Startlist_sheet.cell(row=row_num, column=1, value=rider_name)
            Startlist_sheet.cell(row=row_num, column=i+4, value=rider_results[race_name])

            # count the number of races for the rider
            num_races = sum([1 for cell in Startlist_sheet[row_num][2:] if cell.value == 1])
            Startlist_sheet.cell(row=row_num, column=3, value=num_races)
            riders_sheet.cell(row=row_num, column=4, value=num_races)


            
            # increment the row number for the next rider
            row_num += 1

        # print the race name and the number of riders in the startlist
        print(race_name, ' startlist built')


        #Reset riders and row number
        row_num = 2
        j = 1

# save the workbook
wb.save('riders.xlsx')