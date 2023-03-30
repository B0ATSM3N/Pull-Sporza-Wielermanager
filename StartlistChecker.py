import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import re
import unicodedata

# specify the year to check
year = '2023'

# define the race_names
race_names = [
    "omloop-het-nieuwsblad",
    "kuurne-brussel-kuurne",
    "gp-samyn",
    "strade-bianche",
    "nokere-koers",
    "bredene-koksijde-classic",
    "milano-sanremo",
    "oxyclean-classic-brugge-de-panne",
    "e3-harelbeke",
    "gent-wevelgem",
    "dwars-door-vlaanderen",
    "ronde-van-vlaanderen",
    "scheldeprijs",
    "paris-roubaix",
    "brabantse-pijl",
    "amstel-gold-race",
    "la-fleche-wallone",
    "liege-bastogne-liege"
]

# define the base URL
base_url = 'https://www.procyclingstats.com/'

# load the workbook
wb = load_workbook('riders.xlsx')
riders_sheet = wb['Query result']

## get the Startlist sheet or create a new one if it doesn't exist
if 'Startlist' in wb.sheetnames:
    Startlist_sheet = wb['Startlist']
    Startlist_sheet.delete_rows(2, Startlist_sheet.max_row)
else:
    Startlist_sheet = wb.create_sheet('Startlist')

# write the rider names and race names to the Startlist sheet
Startlist_sheet['A1'] = 'Rider'
Startlist_sheet.column_dimensions['A'].width = 20
Startlist_sheet['B1'] = '# races'
riders_sheet['D1'] = '# races'
for i, race_name in enumerate(race_names):
    Startlist_sheet.cell(row=1, column=i+3, value=race_name)

# initialize row_num to 2
row_num = 2

def clean_name(text):
    # Replace non-ASCII characters with closest ASCII equivalent
    return unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')

# loop through each race in the races sheet and build the startlist
for i, race_name in enumerate(race_names):
        # extract the race name and build the startlist URL
    startlist_url = f'{base_url}race/{str(race_names[i]).lower()}/{year}/startlist'

    # get the HTML content of the startlist URL
    response = requests.get(startlist_url)
    html_content = response.text

    # parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the startlist_v3 element in the HTML content
    startlist_v3 = soup.find('ul', class_='startlist_v3')

    # check if the startlist table was found
    if startlist_v3 is None:
        print('Startlist table not found for race', race_name)
        continue

    startlist_rider_names = set() # create an empty set to store the rider names
    # loop through each team in the startlist_v3 element
    for team in startlist_v3.find_all('li', class_='team'):
        # loop through each rider in the team
        for rider in team.find_all('li'):
            # get the rider name
            startlist_rider_name = clean_name(rider.find_all('span')[-1].text.lower())
            startlist_rider_names.add(startlist_rider_name)

    # loop through each rider in the riders sheet and check if they're in the startlist
    for j, rider in enumerate(riders_sheet.iter_rows(min_row=2, values_only=True)):

        # check if the rider name is not None
        if rider[1] is None:
            continue

        # extract the rider name replace special characters with regular ones
        rider_name = re.sub(r'[^a-zA-Z\s]', '', re.sub(r'[^\w\s.]', '', rider[1]).lower().replace('.', ''))

        rider_result = next(
            (
                1
                for startlist_rider_name in startlist_rider_names
                if rider_name in startlist_rider_name
            ),
            0,
        )
        # write the results for the rider to the Startlist sheet
        Startlist_sheet.cell(row=row_num, column=1, value=rider_name)
        Startlist_sheet.cell(row=row_num, column=i+3, value=rider_result)

        # count the number of races for the rider
        num_races = sum(cell.value == 1 for cell in Startlist_sheet[row_num][2:])
        Startlist_sheet.cell(row=row_num, column=2, value=num_races)
        riders_sheet.cell(row=row_num, column=4, value=num_races)

        row_num += 1

    # print the race name of riders in the startlist
    print(race_name, ' startlist built')


    #Reset riders and row number
    row_num = 2
    j = 1

# save the workbook
wb.save('riders.xlsx')